#-*- coding:utf-8 -*-
'''
    filename: error_code
    version: 2
    improvement: change the python library source file xlutils to openpyxl, because xlutils will destory the excel file
    date: 2019/11/19
'''
import os
import sys
path = os.getcwd() + '/Windows/py3-Tools/3rd'
sys.path.append(path)
path = os.getcwd() + '/Windows/py3-Tools/'
sys.path.append(path)
import re
import openpyxl

#import xlwt
#import xlrd
#reload(sys)
#sys.setdefaultencoding('utf-8')


matchObj_1 = r"^[\t ]*//(.*)"
matchObj_2 = r"^[\t ]*(\w*)[\t ]*="
matchObj_3 = r"^.*=[\t ](-\d+|\d+)"

'''
def GetSheetByName(excelName, sheetName):
    excelName = excelName.decode('utf-8')
    wb = openpyxl.load_workbook(excelName)
    ws = wb.get_sheet_by_name(sheetName)
    return ws
'''
def GetOtherInfoDictByErrorCodeSheet(ws_errorCode):
    max_column = ws_errorCode.max_column
    max_row = ws_errorCode.max_row
    #保存记录信息
    otherInfoDict = {}
    for row in range(6, max_row + 1):
        iID = ws_errorCode.cell(row, 2).value
        otherInfoDict[iID] = []
        for col in range(10, max_column + 1):
            otherInfo = ws_errorCode.cell(row, col).value
            otherInfoDict[iID].append(otherInfo)

    return otherInfoDict


def FillToLanguageSheet(iID, sAnotation, ws_language):
    row = ws_language.max_row + 1
    sNotify = "SYSTEM_ERROR_" + iID
    #print sNotify
    ws_language.cell(row, 1).value = sNotify#.decode('utf-8')
    ws_language.cell(row, 2).value = sAnotation.strip()#.decode('utf-8')

def FillToErrorCodeSheet(sErrorCode, iID, sAnotation, ws_errorCode, otherInfoDict):
    row = ws_errorCode.max_row + 1
    sNotify = "SYSTEM_ERROR_" + iID
    ws_errorCode.cell(row, 1).value = iID#.decode('utf-8')
    ws_errorCode.cell(row, 2).value = sErrorCode 
    ws_errorCode.cell(row, 3).value = "error_titie_system_tips"#.decode('utf-8')
    ws_errorCode.cell(row, 4).value = "系统提示"#.decode('utf-8')
    ws_errorCode.cell(row, 5).value = 2
    ws_errorCode.cell(row, 6).value = sNotify#.decode('utf-8')
    ws_errorCode.cell(row, 7).value = sAnotation.strip()#.decode('utf-8')
    ws_errorCode.cell(row, 8).value = "error_button_haode"
    ws_errorCode.cell(row, 9).value = "好的"#.decode('utf-8')

    try:
        otherList = otherInfoDict[iID]
    except:
        otherList = None

    if otherList != None:
        for i in range(0, len(otherList)):
            col = 9 + i + 1
            ws_errorCode.cell(row, col).value = otherList[i]

def DealSourceFile(sourceFileName, ws_language, ws_errorCode, otherInfoDict):
    # print(sourceFileName)
    with open(sourceFileName, 'r', encoding='utf-8') as fileObj:
        line = fileObj.readline()
        ##直接定位到 table中
        while line:
            if line[0] == "{":
                break
            else:
                line = fileObj.readline()

        line = fileObj.readline()
        sAnotationLine = None
        iCount = 0

        while line:
            sRetCodeObj = re.match(matchObj_2, line)
            iRetCodeObj = re.match(matchObj_3, line)
            if sRetCodeObj and iRetCodeObj:
                iCount = iCount + 1
                sErrorCode = sRetCodeObj.group(1)
                iID = iRetCodeObj.group(1)
                #得到注释
                sAnotationObj = re.match(matchObj_1, sAnotationLine)
                if sAnotationObj:
                    sAnotationLine = ''
                    sAnotation = sAnotationObj.group(1)
                    # print(sAnotation)
                    FillToErrorCodeSheet(sErrorCode, iID, sAnotation, ws_errorCode, otherInfoDict)
                    FillToLanguageSheet(iID, sAnotation, ws_language)
                else:
                    print("ERROR:Parse Fail! ", sErrorCode, iID, " have not Anotation.")
                    sys.exit(-1)
            else:
                sAnotationLine = line
                #print sAnotation.strip()
            line = fileObj.readline()

        #错误码条数
        print(iCount)


def main():
    #languageExcelName = r"yyb_语言包_CN.xlsx".decode('utf-8')
    print(os.getcwd())
    excelPath = os.path.join(os.getcwd(), "DataConfig\Common\wl_网络错误信息表\wl_网络错误信息表.xlsx")
    RetCodePath = os.path.join(os.path.dirname(os.getcwd()), "BuildProtocol\Protocol\cs_retcode.proto")
    #print (excelPath)
    #errorCodeExcelName = "..\..\DataConfig\Common\wl_网络错误信息表\wl_网络错误信息表.xlsx" #.decode('utf-8')
    errorCodeSheetName = "errorcode"
    languageSheetName = "errorcode_lan"
    #sourceFileName = r'..\..\..\BuildProtocol\Protocol\cs_retcode.proto'
    #print(errorCodeExcelName)
    #wb_language = openpyxl.load_workbook(languageExcelName, data_only=True)
    #ws_language = wb_language.get_sheet_by_name(languageSheetName)

    wb_errorCode = openpyxl.load_workbook(excelPath)
    print("Load Excel ")
    ws_errorCode = wb_errorCode[errorCodeSheetName]
    ws_language = wb_errorCode[languageSheetName]

    otherInfoDict = GetOtherInfoDictByErrorCodeSheet(ws_errorCode)

    #删除第六行之后所有行
    ws_errorCode.delete_rows(6, (ws_errorCode.max_row - 4))
    ws_language.delete_rows(6, (ws_language.max_row - 4))

    DealSourceFile(RetCodePath, ws_language, ws_errorCode, otherInfoDict)
    #wb_language.save(languageExcelName)
    wb_errorCode.save(excelPath)



if __name__ == "__main__":
    main()
