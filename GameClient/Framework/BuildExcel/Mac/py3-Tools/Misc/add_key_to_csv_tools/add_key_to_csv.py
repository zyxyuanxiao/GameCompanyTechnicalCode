# encoding : utf-8
import sys
import os

path = os.getcwd() + "/Mac/py3-Tools/"
sys.path.append(path)

path = os.getcwd() + "/Mac/py3-Tools/3rd/"
sys.path.append(path)

path = os.getcwd() + "/Mac/py3-Tools/Ini/"
sys.path.append(path)

from openpyxl import Workbook, load_workbook

import io
import re
import csv
import logging

from Misc.Util import *

# local cfg = config_data_center.getConfigDataByID("dataconfig_activitytime", "ActivityID", activityId)


r_match_key = r'\s*[\"\']\s*(\w+)\s*[\"\']\s*'
r_match_key_value = r'.*'

r_match_key_with_value = r_match_key + ',' + r_match_key_value

match_getConfigDataByID = re.compile(r'getConfigDataByID\(' + r_match_key + r',' + r_match_key_with_value + r'\)')
match_getConfigDatasByID = re.compile(r'getConfigDatasByID\(' + r_match_key + r',' + r_match_key_with_value + r'\)')
match_getConfigDataByTwoID = re.compile(
    r'getConfigDataByTwoID\(' + r_match_key + r',' + r_match_key_with_value + r_match_key_with_value + r'\)')
match_getConfigDatasByTwoID = re.compile(
    r'getConfigDatasByTwoID\(' + r_match_key + r',' + r_match_key_with_value + r_match_key_with_value + r'\)')
match_getConfigDataByThreeID = re.compile(
    r'getConfigDataByThreeID\(' + r_match_key + r',' + r_match_key_with_value + r_match_key_with_value + r_match_key_with_value + r'\)')

r_csharp_config_name = r'\"(\w+)\"'
r_key_name = r'\"(\w+)\"'
csharp_match_getdata = re.compile(r'GetData<.*>\(' + r_csharp_config_name + r'\s*,\s*' + r_key_name + r'.*\)')
csharp_match_getdatas = re.compile(r'GetDatas<.*>\(' + r_csharp_config_name + r'\s*,\s*' + r_key_name + r'.*\)')

math_any_config_name = re.compile(r'dataconfig_(\w+)\"')

CLIENT_KEY_TAG = "_KEY"
NONUNIQUE_TAG = "NON_UNIQUE"


class ConfigKeyInfo:
    def __init__(self, config_name):
        self.__config_name = config_name
        self.__unique_key_list = list()
        self.__multi_key_list = list()

    def AddUniqueKey(self, key_name):
        if not key_name in self.__unique_key_list:
            self.__unique_key_list.append(key_name)

    def AddMultiKey(self, key_name):
        if not key_name in self.__multi_key_list:
            self.__multi_key_list.append(key_name)

    @property
    def ConfigName(self):
        return self.__config_name

    @property
    def UniqueKeyList(self):
        return self.__unique_key_list

    @property
    def MultiKeyList(self):
        return self.__multi_key_list

    def __str__(self):
        result = ''
        result += '    unique_key: ' + str(self.__unique_key_list) + "\n"
        result += '    multi_key: ' + str(self.__multi_key_list) + "\n"
        return result


class ParseClientCode:
    def __init__(self):
        self.config_info = dict()
        self.conflict_config_info = set()
        self.used_by_client = set()
        self.add_new_row = False

    def __Match(self, content, re, unique, match_number):
        matches = re.findall(content)
        if len(matches) > 0:
            for match in matches:
                config_name = match[0]
                config_name = config_name.lower()
                if not config_name in self.config_info:
                    self.config_info[config_name] = ConfigKeyInfo(config_name)
                for i in range(match_number):
                    if unique:
                        self.config_info[config_name].AddUniqueKey(match[i + 1])
                    else:
                        self.config_info[config_name].AddMultiKey(match[i + 1])

    def ParseCSharp(self, csharp_file):
        with open(csharp_file, 'r', encoding='UTF-8-sig') as cs_file:
            content = cs_file.read()
            self.__Match(content, csharp_match_getdata, True, 1)
            self.__Match(content, csharp_match_getdatas, False, 1)

        self.__MatchAny(csharp_file)

    def ParseLuaDir(self, lua_root_path):
        all_files = list(set(GetDirFiles(lua_root_path)))
        all_lua_files = [x for x in all_files if x.split('.')[-1] == 'lua']

        for lua_file_path in all_lua_files:
            self.__ParseLuaFile(lua_file_path)
            self.__MatchAny(lua_file_path)

    def GenConflict(self):
        for config_name, info in self.config_info.items():
            for key in info.UniqueKeyList:
                if key in info.MultiKeyList:
                    self.conflict_config_info.add(config_name)
                    break

    def __MatchAny(self, path):
        with open(path, 'r', encoding='UTF-8-sig') as f:
            content = f.read()
            matches = math_any_config_name.findall(content)
            if len(matches) > 0:
                for match in matches:
                    config_name = match.lower()
                    self.used_by_client.add(config_name)

    def __ParseLuaFile(self, lua_file_path):
        with open(lua_file_path, 'r', encoding='UTF-8-sig') as lua_file:
            content = lua_file.read()

            self.__Match(content, match_getConfigDataByID, True, 1)
            self.__Match(content, match_getConfigDatasByID, False, 1)
            self.__Match(content, match_getConfigDataByTwoID, False, 2)
            self.__Match(content, match_getConfigDatasByTwoID, False, 2)
            self.__Match(content, match_getConfigDataByThreeID, False, 3)

    def GetConfigMultiKey(self):
        path = os.getcwd() + "/Data/ConfigMultiKey.csv"
        all_rows = []
        with open(path, 'r', encoding='UTF-8-sig') as csv_file:
            csv_reader = csv.reader(csv_file)
            for row in csv_reader:
                clist = row[1].split(";")
                for c in clist:
                    if c == '':
                        clist.remove(c)
                all_rows.append({row[0]: clist})

        return all_rows

    def add_first_row_csv(self, path):
        # if not self.add_new_row:
        #     return
        all_rows = []
        MK = self.GetConfigMultiKey()
        with open(path, newline='', encoding='UTF-8-sig') as csv_file:
            csv_reader = csv.reader(csv_file)
            first_row = next(csv_reader)
            # 服务器的不管添加 key 列
            if first_row[0].lower() == "s":
                return

            for cmk in MK:
                for k, v in cmk.items():
                    if path.find(k) > 0:
                        for index in v:
                            first_row[int(index)] = first_row[int(index)] + CLIENT_KEY_TAG + "\n" + NONUNIQUE_TAG
            if first_row[0].find(CLIENT_KEY_TAG) <= 0:
                first_row[0] = first_row[0] + CLIENT_KEY_TAG

            for row in csv_reader:
                all_rows.append(row)
            all_rows.insert(0, first_row)

        with open(path, 'w', newline='', encoding='UTF-8-sig') as csv_file:
            csv_writer = csv.writer(csv_file)
            csv_writer.writerows(all_rows)

    # def AddSixRow_xlsx(self, sheet):
    #     if not self.add_new_row:
    #         return
    #     sheet.insert_rows(6)

    def Add_Key_And_NON_UNIQUE_To_CSV(self, csv_path):
        all_files = list(set(GetDirFiles(csv_path)))

        # 对 csv 进行 关键key添加
        all_csv_files = [x for x in all_files if x.split('.')[-1] == 'csv']
        for csv_file_path in all_csv_files:
            # 不让 Excel 的 缓存文件进入监测表格的代码里面
            if "$" in csv_file_path:
                continue
            if "~" in csv_file_path:
                continue
            if "@" in csv_file_path:
                continue
            if "#" in csv_file_path:
                continue
            self.add_first_row_csv(csv_file_path)
            print("给  " + csv_file_path + "  添加 key完成")

            # config_file_name = csv_file_path.split('/')[-1].split('.')[-2]
            # config_file_name = 'dataconfig_' + config_file_name.lower()
            # if config_file_name in self.config_info and not (config_file_name in self.conflict_config_info):
            #     key_info = self.config_info[config_file_name]
            #     try:
            #         self.__WriteKeyInfoToCsv(csv_file_path, key_info)
            #     except Exception as e:
            #         print(e)

        # 目前的方案是不对xlsx就行 key 值添加
        # 对 xlsx进行 关键key添加
        # all_xlsx_files = [x for x in all_files if x.split('.')[-1] == 'xlsx']
        # with open('xlsx_files.txt', 'w+') as f:
        #     for xlsx_file_path in all_xlsx_files:
        #         xlsx_file_path = xlsx_file_path.replace('\\', '/')
        #         file_name = xlsx_file_path.split('/')[-1]
        #
        #         wb = load_workbook(filename = xlsx_file_path)
        #         dirty = False
        #         for sheet_name in wb.sheetnames:
        #             self.AddSixRow_xlsx(wb[sheet_name])
        #             config_name = 'dataconfig_' + sheet_name.lower()
        #             if config_name in self.config_info and config_name not in self.conflict_config_info:
        #                 dirty = self.__WriteKeyInfoToXlsx(wb[sheet_name], self.config_info[config_name], xlsx_file_path)
        #
        #         try:
        #             if dirty:
        #                 f.write('{}\n'.format(xlsx_file_path))
        #                 wb.save(xlsx_file_path)
        #         except Exception as e:
        #             print(xlsx_file_path)
        #             print(e)
        #             pass

    # def __WriteKeyInfoToCsv(self, csv_file_path, key_info):
    #     unique_list = key_info.UniqueKeyList
    #     multi_key_list = key_info.MultiKeyList
    #     all_rows = []
    #     unique_row_nums, multi_row_nums = [], []
    #     with open(csv_file_path, newline='', encoding='GB18030') as csv_file:
    #         csv_reader = csv.reader(csv_file)
    #         row_num = 0
    #         for row in csv_reader:
    #             all_rows.append(row)
    #             if row_num == 3:
    #                 for i in range(len(row)):
    #                     if row[i].strip() in unique_list:
    #                         unique_row_nums.append(i)
    #                     if row[i].strip() in multi_key_list:
    #                         multi_row_nums.append(i)
    #
    #             if row_num == 4:
    #                 for unique_row in unique_row_nums:
    #                     row[unique_row] = row[unique_row].replace(CLIENT_KEY_TAG, '')
    #
    #                 for multi_row in multi_row_nums:
    #                     row[multi_row] = row[multi_row].replace(CLIENT_KEY_TAG + NONUNIQUE_TAG, '')
    #
    #             if row_num == 5:
    #                 for unique_row in unique_row_nums:
    #                     if CLIENT_KEY_TAG not in row[unique_row]:
    #                         row[unique_row] = CLIENT_KEY_TAG
    #
    #                 for multi_row in multi_row_nums:
    #                     if NONUNIQUE_TAG not in row[multi_row]:
    #                         row[multi_row] = CLIENT_KEY_TAG + NONUNIQUE_TAG
    #
    #             row_num += 1
    #
    #     with open(csv_file_path, 'w', newline='', encoding='GB18030') as csv_file:
    #         csv_writer = csv.writer(csv_file)
    #         csv_writer.writerows(all_rows)

    # def __WriteKeyInfoToXlsx(self, sheet, key_info, file):
    #     col_count = len(sheet[4])
    #     fourth_row, sixth_row = sheet[4], sheet[6]
    #     dirty = False
    #     for i in range(col_count):
    #         attribute_name = str(fourth_row[i].value).strip()
    #         comment = str(sixth_row[i].value)
    #         if attribute_name != None and attribute_name in key_info.UniqueKeyList and (CLIENT_KEY_TAG.strip() not in comment and CLIENT_KEY_TAG != comment):
    #             dirty = True
    #             if comment != 'None':
    #                 print(1,  comment, file)
    #                 continue
    #             print(1, sixth_row[i].value, file)
    #             sixth_row[i].value = CLIENT_KEY_TAG
    #
    #         if attribute_name != None and attribute_name in key_info.MultiKeyList and NONUNIQUE_TAG not in comment:
    #             dirty = True
    #             if comment != 'None':
    #                 continue
    #                 print(2, comment, file)
    #             print(2, sixth_row[i].value, file)
    #             sixth_row[i].value = CLIENT_KEY_TAG + NONUNIQUE_TAG
    #     return dirty

    def GenLog(self):
        with open("Mac/Log/conflict_key_info_output.txt", 'w+') as file:
            for config_name, info in self.config_info.items():
                if config_name in self.conflict_config_info:
                    file.write(config_name + '\n')
                    file.write(str(info))

        with open("Mac/Log/parse_csv_key_info_output.txt", 'w+') as file:
            for config_name, info in self.config_info.items():
                if config_name in self.conflict_config_info:
                    continue
                file.write(config_name + '\n')
                file.write(str(info))

    def __GetExportList(self, path):
        exported_by_md5config = list()
        with open(path) as f:
            r = csv.reader(f)
            next(r)
            for row in r:
                if len(row[2].strip()) > 0 or len(row[3].strip()) > 0:
                    exported_by_md5config.append(row[0])
        return exported_by_md5config

    def GenNotUsedByClientLog(self):
        p1Path = os.getcwd() + "/Data/Md5_Common(cehua_China).csv"
        p2Path = os.getcwd() + "/Data/MD5_yunying/MD5yunying_China.csv"
        p1 = self.__GetExportList(p1Path)
        p2 = self.__GetExportList(p2Path)
        final_p = p1 + p2

        with open("Mac/Log/not_used_by_client.txt", 'w+') as f:
            for path in final_p:
                config_name = path.split('\\')[-1].split('.')[0].lower()
                if config_name not in self.used_by_client:
                    f.write('{}\n'.format(path))


def add_key_to_all_csv():
    # luaPath = os.getcwd() + "/Data/Client/lua/"
    # csPath = os.getcwd() + "/Data/Client/CSResBinData/ResBinData.cs"
    tool = ParseClientCode()
    # tool.ParseLuaDir(luaPath)
    # tool.ParseCSharp(csPath)
    # tool.GenConflict()
    # tool.GenLog()
    # tool.GenNotUsedByClientLog()  # 不在这个地方对 csv 的 md5 进行检查
    csv_path = os.getcwd() + "/DataConfig/"
    tool.Add_Key_And_NON_UNIQUE_To_CSV(csv_path)


def add_key_to_one_csv():
    # luaPath = os.getcwd() + "/Data/Client/lua/"
    # csPath = os.getcwd() + "/Data/Client/CSResBinData/ResBinData.cs"
    tool = ParseClientCode()
    # tool.ParseLuaDir(luaPath)
    # tool.ParseCSharp(csPath)
    # tool.GenConflict()
    # tool.GenLog()
    # tool.GenNotUsedByClientLog() #不在这个地方对 csv 的 md5 进行检查
    csvPath = sys.argv[1]
    tool.Add_Key_And_NON_UNIQUE_To_CSV(csvPath)


if __name__ == "__main__":
    if len(sys.argv) >= 2 and os.getcwd() in sys.argv[1]:
        add_key_to_one_csv()
    else:
        add_key_to_all_csv()
    print("添加 key 结束=======================================================")
