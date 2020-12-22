# coding=utf-8
from openpyxl import load_workbook
import sys
import os

path = os.getcwd() + '/Mac/py3-Tools/3rd'
sys.path.append(path)


def CheckXlsxSheetName(excelPath, keyExcelPath):
    print(excelPath)
    workbook = load_workbook(excelPath, data_only=True)
    for name in workbook.sheetnames:
        print(name)
    workbook.close()
    print("=======================================")


def GetAllExcel(excelPath, keyExcelPath):
    keyExcelPath = str(keyExcelPath).replace("\\", "/")
    keyExcelPath = str(keyExcelPath).replace("//", "/")
    keyExcelPaths = str(keyExcelPath).split("/")
    keyExcelPath = keyExcelPaths[len(keyExcelPaths) - 1]

    list_dirs = os.walk(excelPath)
    for root, dirs, files in list_dirs:
        for f in files:
            if ".xlsx" not in f:
                continue
            if "$" in f:
                continue
            if "~" in f:
                continue
            if "@" in f:
                continue
            if "#" in f:
                continue
            CheckXlsxSheetName(root + "/" + f, keyExcelPath)


if __name__ == "__main__":
    list_sheet_name = []
    EXCEL_PATH = os.path.dirname(os.path.dirname(
        os.path.dirname(os.path.dirname(__file__)))) + "/DataConfig/"
    GetAllExcel(EXCEL_PATH, "xlsx")
