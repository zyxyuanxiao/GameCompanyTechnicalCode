# coding=utf-8
import sys
import os

path = os.getcwd() + '/Mac/py3-Tools/3rd'
sys.path.append(path)

# 打开文件：
from openpyxl import load_workbook
from tkinter import messagebox

'''
检查所有的 Excel 表中的 sheet 名字是否有重复的
检查所有的 Excel 表中的 sheet 中的列是否有重复的,以及头四行是否是正常的标准的写法
'''

'''
前三行是固定值
第一行某个值可以为空,C 表示只有客户端读,CS 表示客户端和服务器都读,S 表示只有服务器读,""表示是备注,或者没有数据

第二行某个值可以为空,
optional 表示可填写可不填写字段,
required表示必填字段,
repeated表示重复字段,optional_struct表示重复字段的子字段,
""表示是备注,或者没有数据

第三行某个值可以为空,
bool表示布尔值,True 或 False,
uint32      无符号整型,32 位
sint32      有符号整型,32 位
int32       无符号整型,32 位
uint64      无符号整型,64 位
sint64      有符号整型,64 位
int64/int   无符号整型,64 位
float       浮点数类型

第四行可以是英文,汉字,还有空字符.
'''
Excel_First_Row = ["C", "CS", "S", ""]
Excel_Second_Row = ["optional", "", "required", "repeated", "optional_struct"]
Excel_Third_Row = ["bool", "", "uint32", "sint32", "int32", "uint64", "sint64", "int64", "int", "float", "string"]

list_sheet_name = []


# 是否包含中文
def is_contains_chinese(strs):
    for _char in strs:
        if '\u4e00' <= _char <= '\u9fa5':
            return True
    return False


# 根据路径,得到所有的 Excel
def GetAllExcel(excelPath, keyExcelPath):
    keyExcelPath = str(keyExcelPath).replace("\\", "/")
    keyExcelPath = str(keyExcelPath).replace("//", "/")
    keyExcelPaths = str(keyExcelPath).split("/")
    keyExcelPath = keyExcelPaths[len(keyExcelPaths) - 1]

    list_dirs = os.walk(excelPath)
    for root, dirs, files in list_dirs:
        for f in files:
            if ".xlsx" not in f:
                continue
            if "$" in f:
                continue
            if "~" in f:
                continue
            if "@" in f:
                continue
            if "#" in f:
                continue
            CheckXlsxSheetName(root + "/" + f, keyExcelPath)


# 是否是数字
def is_number(s):
    try:  # 如果能运行float(s)语句，返回True（字符串s是浮点数）
        float(s)
        return True
    except ValueError:  # ValueError为Python的一种标准异常，表示"传入无效的参数"
        pass  # 如果引发了ValueError这种异常，不做任何事情（pass：不做任何事情，一般用做占位语句）
    try:
        import unicodedata  # 处理ASCii码的包
        for i in s:
            unicodedata.numeric(i)  # 把一个表示数字的字符串转换为浮点数返回的函数
            # return True
        return True
    except (TypeError, ValueError):
        pass
    return False


regexString = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ_123456789"


def is_field(stringField):
    for char in stringField:
        if char not in regexString:
            return False
    return True


def CheckXlsxSheetName(excelPath, keyExcelPath):
    # print(excelPath )
    workbook = load_workbook(excelPath, data_only=True)
    for name in workbook.sheetnames:
        sheet_table = workbook[name]
        # 如果sheet名是//开头不转化成csv
        if sheet_table.title.startswith('#'):
            continue
        # 如果sheet名是包含sheet字段的不给打表
        if sheet_table.title.lower().find("sheet") >= 0:
            continue
        # 如果sheet名是汉字开头不转化成csv
        if is_contains_chinese(sheet_table.title):
            continue
        # 如果当前的 Excel 表中的数据为空,或者前 6 行没有数据,前 2 列没有数据,则不能进行转表
        one_cell = str(sheet_table.cell(row=1, column=1).value).strip()
        if sheet_table.max_row < 7 or sheet_table.max_column < 2 or (one_cell not in Excel_First_Row):
            continue

        # 检查重名操作
        for data in list_sheet_name:
            if sheet_table.title in data.keys():
                messagebox.showerror(sheet_table.title, " 重名了,请先删除或者修改再打表,路径  " + data[sheet_table.title])
                messagebox.showerror(sheet_table.title, " 重名了,请先删除或者修改再打表,路径  " + excelPath)

        list_sheet_name.append({sheet_table.title: excelPath})
        # 单独对当前打表的 xlsx 里面的数据进行检查
        if keyExcelPath in excelPath:
            CheckXlsxSheetContent(sheet_table, excelPath)
    workbook.close()


def CheckXlsxSheetContent(sheet_table, excelPath):
    all_key_names = []
    for col_num in range(1, sheet_table.max_column + 1):
        # 判断第一行的数据
        row_data1 = str(sheet_table.cell(row=1, column=col_num).value)
        row_data1 = "" if row_data1 == "None" else row_data1
        # print("row_data1:" + row_data1)
        if row_data1 not in Excel_First_Row:
            messagebox.showerror(sheet_table.title,
                                 excelPath + "   " + "   data:" + row_data1 + "  格式不正确了,请检查第 1 行,第 " + str(
                                     col_num) + " 列")

        # 判断第二行的数据
        row_data2 = str(sheet_table.cell(row=2, column=col_num).value)
        row_data2 = "" if row_data2 == "None" else row_data2
        # print("row_data2:" + row_data2)
        if row_data2 not in Excel_Second_Row:
            messagebox.showerror(sheet_table.title,
                                 excelPath + "   " + "   data:" + row_data2 + " 格式不正确,请检查第 2 行,第 " + str(
                                     col_num) + " 列")

        # 判断第三行的数据
        row_data3 = str(sheet_table.cell(row=3, column=col_num).value)
        row_data3 = "" if row_data3 == "None" else row_data3
        # print("row_data3:" + row_data3)
        if (not is_number(row_data3)) and (row_data3 not in Excel_Third_Row):
            messagebox.showerror(sheet_table.title,
                                 excelPath + "   " + "   data:" + row_data3 + " 格式不正确了,请检查第 3 行,第 " + str(
                                     col_num) + " 列")

        # 判断第四行的数据,如果包含汉字,则检查第二行是否有数据,没有数据,则跳过,如果有数据,则不能为空
        row_data4 = str(sheet_table.cell(row=4, column=col_num).value)
        row_data4 = "" if row_data4 == "None" else row_data4
        # print("row_data4:" + row_data4)
        if is_contains_chinese(row_data4) or row_data2 == "":
            continue
        elif (row_data4 != "") and (not is_field(row_data4)):
            messagebox.showerror(sheet_table.title,
                                 excelPath + "   " + "   data:" + row_data4 + " 格式不正确,请检查第 4 行,第 " + str(
                                     col_num) + " 列")

        row_data4 = row_data4.lower()
        if (is_number(row_data3)) and row_data2 == "repeated":
            if row_data4 == "":
                row_data4 = str(col_num) + "__repeated__"
            row_data4 = {row_data4: row_data3}
        if (is_number(row_data3)) and row_data2 == "optional_struct":
            if row_data4 == "":
                messagebox.showerror(sheet_table.title,
                                     excelPath + "   data:" + row_data4 + " 格式不正确,请检查第 4 行,第 " + str(
                                         col_num) + " 列,optional_struct下方字段名不应该为空")
            row_data4 = {row_data4: row_data3}
        all_key_names.append(row_data4)

    all_key_names = list(filter(None, all_key_names))
    # 记录repeated以及optional_struct之后的参数数据
    repeat_key_range = {}
    for i in range(len(all_key_names)):
        data = all_key_names[i]
        if isinstance(data, dict):
            for k, v in data.items():
                nextData = all_key_names[i + 1]
                if isinstance(nextData, dict):
                    all_key_names[i] = k
                    for k1, v1 in nextData.items():
                        length = int(float(v)) * int(float(v1))
                        repeat_key_range[i + 1 + 1] = i + 1 + 1 + length
                        all_key_names[i + 1] = k1

    for i in sorted(repeat_key_range.keys(), reverse=True):
        k, v = (i, repeat_key_range[i])
        del all_key_names[k:v]
    # 检查是否有重复的值
    not_repeated_list = set(all_key_names)
    if len(all_key_names) != len(not_repeated_list):
        for not_repeated in not_repeated_list:
            all_key_names.remove(not_repeated)
        messagebox.showerror("有重复的字段值,请检查 ", excelPath + "    " + sheet_table.title + "   字段为:" + str(all_key_names))


# def CheckAll(keyExcelPath):
#     print("开始检查配置表:\n", keyExcelPath)
#     global list_sheet_name
#     list_sheet_name = []
#     GetAllExcel(GetAllExcelPath(), keyExcelPath)
#     print("检查完毕配置表,无错误")


# 获取所有的表的名字
def GetAllExcelPath():
    # 测试与正式的路径
    EXCEL_PATH = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))) + "/DataConfig/"
    # EXCEL_PATH = "/Users/xlcw/Work/tw/Turing_iOS/BuildDataConfig/DataConfig"
    # EXCEL_PATH = "/Users/xlcw/Work/tw/Turing_iOS/BuildDataConfig/DataConfig/Common/ai_配置信息表"
    # EXCEL_PATH = "/Users/xlcw/Work/tw/Turing_iOS/BuildDataConfig/DataConfig/Common/fl_福利中心"
    # EXCEL_PATH = "/Users/xlcw/Work/tw/Turing_iOS/BuildDataConfig/DataConfig/Common/服务器_机器人数据"
    # EXCEL_PATH = "/Users/xlcw/Work/tw/Turing_iOS/BuildDataConfig/DataConfig/Common/lt_聊天配置表"
    return EXCEL_PATH


if __name__ == "__main__":
    print("开始检查配置表,觉得打表一定没有问题的可以关闭")
    list_sheet_name = []
    GetAllExcel(GetAllExcelPath(), "xlsx")
    print("检查完毕配置表,无错误")
